from openpyxl import  Workbook, load_workbook
from tkinter import *
from customtkinter import *
import tkinter
from PIL import Image
from tkinter.font import Font
import rows_def
number = 1
s = ''

root = CTk()

path_scripts = r"C:\Users\Vasya\PycharmProjects\pythonProject5\script_things\scripts.txt"

path_try = r"C:\Users\Vasya\PycharmProjects\pythonProject5\script_things\try.txt"


wb = load_workbook(r'C:\Users\Vasya\PycharmProjects\pythonProject5\exel\auto2.xlsx')
ws = wb.active

def nextA(n1):
    search_text = '111'
    n = 'A'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)


def nextB(n1):
    search_text = '222'
    n = 'B'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextC(n1):
    search_text = '333'
    n = 'C'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextD(n1):
    search_text = '444'
    n = 'D'
    r = n + str(n1)
    replace_text = str(ws[r].value)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextG(n1):
    search_text = '555'
    n = 'G'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextH(n1):
    search_text = '666'
    n = 'H'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextI(n1):
    search_text = '777'
    n = 'I'
    r = n + str(n1)
    replace_text = str(ws[r].value)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextJ(n1):
    search_text = '888'
    n = 'J'
    r = n + str(n1)
    replace_text = str(ws[r].value)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextK(n1):
    search_text = '999'
    n = 'K'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextL(n1):
    search_text = '1010'
    n = 'L'
    r = n+str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextM(n1):
    search_text = '1111'
    n = 'M'
    r = n+str(n1)
    replace_text = str(ws[r].value)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)


def nextN(n1):
    search_text = '1212'
    n = 'N'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def drp_dwn(n1):
    search_text = '1313'
    n = 'E'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Wybierz":
            replace_text = 'op0'
        if replace_text == "Podlaskie":
            replace_text = 'op1'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def drp_dwn_2(n1):
    search_text = '1414'
    n = 'F'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Wybierz":
            replace_text = 'op0'
        if replace_text == "Miasto Suwałki":
            replace_text = 'op1'
        if replace_text == "suwalski":
            replace_text = 'op2'
        if replace_text == "sejneński":
            replace_text = 'op3'
        if replace_text == "sejneński":
            replace_text = 'op4'
        if replace_text == "grajewski":
            replace_text = 'op5'
        if replace_text == "moniecki":
            replace_text = 'op6'
        if replace_text == "sokólski":
            replace_text = 'op7'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_1(n1):
    search_text = '1515'
    n = 'O'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "niższe niż podstawowe":
            replace_text = 'op1'
        if replace_text == "podstawowe":
            replace_text = 'op2'
        if replace_text == "gimnazjalne":
            replace_text = 'op3'
        if replace_text == "ponadgimnazjalne":
            replace_text = 'op4'
        if replace_text == "policealne":
            replace_text = 'op5'
        if replace_text == "wyższe":
            replace_text = 'op6'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_2(n1):
    search_text = '1616'
    n = 'P'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "osoba bezrobotna":
            replace_text = 'op1'
        if replace_text == "osoba bierna zawodowo":
            replace_text = 'op2'
        if replace_text == "osoba pracująca":
            replace_text = 'op3'
        if replace_text == "osoba prowadząca działalność na własny rachunek":
            replace_text = 'op4'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_3(n1):
    search_text = '1717'
    n = 'Q'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_4(n1):
    search_text = '1818'
    n = 'R'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_5(n1):
    search_text = '1919'
    n = 'S'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_6(n1):
    search_text = '2020'
    n = 'T'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_7(n1):
    search_text = '2121'
    n = 'U'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_8(n1):
    search_text = '2t2t'
    n = 'V'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)

def radio_9(n1):
    search_text = '2323'
    n = 'W'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_10(n1):
    search_text = '2424'
    n = 'AA'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def radio_11(n1):
    search_text = '2525'
    n = 'AB'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)


def radio_12(n1):
    search_text = '2626'
    n = 'AC'
    r = n + str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        if replace_text == "Tak":
            replace_text = 'op1'
        if replace_text == "Nie":
            replace_text = 'op2'
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextAD(n1):
    search_text = '!!!'
    n = 'AD'
    r = n+str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)

def nextAE(n1):
    search_text = '!!!!'
    n = 'AE'
    r = n+str(n1)
    replace_text = ws[r].value
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(search_text, replace_text)
    with open(path_scripts, 'w') as f:
        f.write(data)
    with open(path_scripts, 'r') as f:
        data = f.read()
        data = data.replace(replace_text, search_text)
label = CTkLabel(root, text=number)
label.pack()

def g_next():
    with open(r'C:\Users\Vasya\PycharmProjects\pythonProject5\script_things\try.txt','r') as f:
        with open(r'C:\Users\Vasya\PycharmProjects\pythonProject5\script_things\scripts.txt','w') as f1:
            for line in f:
                f1.write(line)
    global number
    number += 1
    label.configure(text=number)
    rows_def.nextM(number)
    rows_def.nextA(number)
    rows_def.nextB(number)
    rows_def.nextC(number)
    rows_def.nextD(number)
    rows_def.nextG(number)
    rows_def.nextH(number)
    rows_def.nextI(number)
    rows_def.nextJ(number)
    rows_def.nextK(number)
    rows_def.nextL(number)
    rows_def.nextN(number)
    rows_def.drp_dwn(number)
    rows_def.drp_dwn_2(number)
    rows_def.radio_1(number)
    rows_def.radio_2(number)
    rows_def.radio_3(number)
    rows_def.radio_4(number)
    rows_def.radio_5(number)
    rows_def.radio_6(number)
    rows_def.radio_7(number)
    rows_def.radio_8(number)
    rows_def.radio_9(number)
    rows_def.radio_10(number)
    rows_def.radio_11(number)
    rows_def.radio_12(number)
    rows_def.nextAE(number)
    rows_def.nextAD(number)
    f = open(path_scripts, 'r')
    c = f.read()
    root.clipboard_clear()
    root.clipboard_append(c)
    print("Number of row: ",number)
    s = "Number of row: " + str(number)
    


def reset_script():
    global number
    number = 1
    label.configure(text=number)
    